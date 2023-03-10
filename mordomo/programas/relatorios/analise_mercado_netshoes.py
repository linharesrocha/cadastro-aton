from bs4 import BeautifulSoup
import pandas as pd
import requests
import os
import re
import io
from PIL import Image
import win32com.client as win32
from time import sleep
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

os.system('cls')
pesquisa = input('\nPesquisa Netshoes: ')

url_mais_populares = f'https://www.netshoes.com.br/busca?nsCat=Natural&q={pesquisa}'
url_mais_vendidos = f'https://www.netshoes.com.br/busca?nsCat=Natural&q={pesquisa}&sort=best-sellers'
url_lancamentos = f'https://www.netshoes.com.br/busca?nsCat=Natural&q={pesquisa}&sort=new-releases'
url_ofertas = f'https://www.netshoes.com.br/busca?nsCat=Natural&q={pesquisa}&sort=offers'
url_mais_avaliados = f'https://www.netshoes.com.br/busca?nsCat=Natural&q={pesquisa}&sort=review-stars'
url_maior_preco = f'https://www.netshoes.com.br/busca?nsCat=Natural&q={pesquisa}&sort=highest-first'
url_menor_preco = f'https://www.netshoes.com.br/busca?nsCat=Natural&q={pesquisa}&sort=lowest-first'

print(f'\nOrdenar por:\n1. Mais Populares\n2. Mais Vendidos\n3. Lançamentos\n4. Ofertas\n5. Melhor Avaliados\n6. Maior Preço\n7. Menor Preço\n')

while True:
    opcao = str(input('Digite uma opção: '))
    if re.match('^[1-7]$', opcao):
        opcao = int(opcao)
        if opcao >= 1 and opcao <= 7:
            if opcao == 1:
                url = url_mais_populares
            elif opcao == 2:
                url = url_mais_vendidos
            elif opcao == 3:
                url = url_lancamentos
            elif opcao == 4:
                url = url_ofertas
            elif opcao == 5:
                url = url_mais_avaliados
            elif opcao == 6:
                url = url_maior_preco
            elif opcao == 7:
                url = url_menor_preco
            
            break

print(' ')
while True:
    quantidade_anuncios = input('Quantos anúncios? (max 42): ')
    if quantidade_anuncios.isdigit():
        quantidade_anuncios = int(quantidade_anuncios)
        if quantidade_anuncios > 0 and quantidade_anuncios <= 42:
            break


user_agent = {'User-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko)' 'Chrome/106.0.0.0 Safari/537.36'}

wb = Workbook()
ws = wb.active

titulo_list = []
preco_de_list = []
preco_por_list = []
avaliacoes_list = []
notas_list = []
link_anuncios_list = []
posicao_list = []
link_photos_list = []

page = requests.get(url, headers=user_agent)
site = BeautifulSoup(page.content, "html.parser")

# Obtendo links
container = site.find(class_="item-list")

# Coletando todos links href do atributo a
all_links = [link['href'] for link in container.find_all('a')]

# Verificando quantos anúncios deu
if len(all_links) < int(quantidade_anuncios):
    print(f'\nATENÇÃO!\nCom essa busca, retornou apenas {len(all_links)} resultados!\n')
    sleep(5)
    
# Filtrando apenas os links que começam com //www.netshoes.com
netshoes_links = [link.replace('//','https://') for link in all_links if link.startswith('//www.netshoes.com')]

# Removendo links que não são de anúncios
for i in range(len(netshoes_links)-1, -1, -1):
    if "//www.netshoes.com.br/busca?" in netshoes_links[i]:
        del netshoes_links[i]

# Entrando em cada link (produto)
os.system('cls')
for index, link in enumerate(netshoes_links):
    
    if index == quantidade_anuncios:
        break
    
    print(f'{str(index+1)}/{str(quantidade_anuncios)}')
    
    page = requests.get(link, headers=user_agent)
    site = BeautifulSoup(page.content, "html.parser")
    
    container = site.find(id='content')
    
    # Posição do anúncio
    posicao_list.append(index+1)
    
    # Titulo do Anúncio
    title = container.find(class_="short-description").find('h1').getText()
    
    # Nem sempre terá Preço DE
    try:
        preco_de = container.find(class_="reduce").getText().replace('R$ ', '')
    except:
        preco_de = '0'
        
    # Preço Por do Anúncio
    preco_por = container.find('div', class_="default-price").find('strong').getText().replace('R$ ', '')
    
    # Avaliações do Produto
    try:
        container_review = container.find(class_="rating-box")
        avaliacoes = container_review.find('span', class_='rating-box__numberOfReviews').getText().replace(' avaliações', '').replace(' avaliação', '')
        nota = container_review.find(class_='rating-box__value').getText()
    except:
        avaliacoes = '0'
        nota = '0'
        
    # Foto
    link_photo = container.find('img', class_='zoom').get('src')
    
    
    # Append lista
    titulo_list.append(title)
    preco_de_list.append(preco_de)
    preco_por_list.append(preco_por)
    avaliacoes_list.append(avaliacoes)
    notas_list.append(nota)
    link_anuncios_list.append(link)
    link_photos_list.append(link_photo)
    
    os.system('cls')
    print(f'\nPesquisando {pesquisa} na Netshoes...\n')

    

# Criar o dicionário
dicionario = {
    "posicao": posicao_list,
    "titulo": titulo_list,
    "preco_de": preco_de_list,
    "preco_por": preco_por_list,
    "avaliacoes": avaliacoes_list,
    "notas": notas_list,
    "link_fotos": link_photos_list,
    "link_anuncios": link_anuncios_list
}

# Criar o dataframe
df = pd.DataFrame(dicionario)

# Converte colunas para float e int
df['preco_de'] = df['preco_de'].str.replace(',','.').astype(float)
df['preco_por'] = df['preco_por'].str.replace(',','.').astype(float)
df['notas'] = df['notas'].str.replace(',','.').astype(float)
df['avaliacoes'] = df['avaliacoes'].astype(int)

# remover substring logo depois da substring '.jpg'
df['link_fotos'] = df['link_fotos'].str.replace(r'\.jpg.*', '.jpg', regex=True)


# Carregue o arquivo Excel
workbook = Workbook()

# Selecione a planilha que contém os dados
worksheet = workbook.active

# Copie os dados do dataframe para a planilha
for r in dataframe_to_rows(df, index=False, header=True):
    worksheet.append(r)

os.system('cls')
print('\nCarregando os dados da planilha...\n')

# Percorra a coluna contendo os links de imagem
for row in worksheet.iter_rows(min_row=2, min_col=7, max_col=7):
    for cell in row:
        # Obtenha o link da imagem da célula
        link = cell.value
        # Se o link estiver vazio, pule para a próxima célula
        if not link:
            continue
        # Baixe a imagem do link usando requests
        response = requests.get(link)
        # Verifique se o download foi bem-sucedido
        if response.status_code != 200:
            print(f"Erro ao baixar a imagem: {response.status_code}")
            continue
        # Crie um objeto Image a partir do conteúdo da imagem
        image = Image(io.BytesIO(response.content))
        # Ajuste o tamanho da imagem para caber na célula
        column_width = worksheet.column_dimensions[cell.column_letter].width
        if column_width is None:
            column_width = 10 # Largura padrão de 10 para colunas não definidas
        row_height = worksheet.row_dimensions[cell.row].height
        if row_height is None:
            row_height = 20 # Altura padrão de 20 para linhas não definidas
        image.width = 8 * (column_width - 2)
        image.height = 7 * (row_height - 2)
        # Adicione a imagem à célula e ajuste o estilo da célula
        worksheet.add_image(image, cell.coordinate)
        cell.alignment = Alignment(wrapText=True, vertical='top')
        cell.value = None

# Adiicona um filtro na primeira linha
worksheet.auto_filter.ref = 'A1:H1'

# ajusta a largura da coluna B para o tamanho máximo do conteúdo
worksheet.column_dimensions['B'].width = 80

# Aumentando o tamanho das linhas
# itera sobre todas as linhas e define a altura desejada
aux = 0
for row in worksheet.rows:
    # Pula o cabeçalho
    if aux != 0:
        row_height = 97.5
        row_dimensions = worksheet.row_dimensions[row[0].row]
        row_dimensions.height = row_height
    aux = aux + 1

# Salve as alterações no arquivo Excel
pesquisa = pesquisa.replace(' ', '-')
nome_arquivo = f'pesquisa-netshoes-{pesquisa}.xlsx'
pasta_downloads = os.path.expanduser('~\Downloads')

# Verifica se o arquivo já existe na pasta de downloads
caminho_arquivo = os.path.join(pasta_downloads, nome_arquivo)
contador = 1

while os.path.exists(caminho_arquivo):
    # Se o arquivo já existe, acrescenta um número ao nome do arquivo
    nome_arquivo_modificado = f"{contador}-{nome_arquivo}"
    caminho_arquivo = os.path.join(pasta_downloads, nome_arquivo_modificado)
    contador += 1
os.system('cls')

dados_ws = workbook.active
if dados_ws.max_row > 1:
    
    # Salvando o arquivo
    workbook.save(f'{caminho_arquivo}')
    sleep(1)
    
    # Abre o excel
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Workbooks.Open(caminho_arquivo)
    excel.Visible = True
    
    print(f'A PLANILHA FOI ABERTA! SIGA AS INSTRUÇÕES ABAIXO.\n\n1. Exclua as linhas dos produtos errados.\n2. Salve as alterações.\n3. Feche o arquivo excel!\n')
    while True:
        awnser = input('Digite OK para continuar: ').upper()
        if awnser == 'OK':
            break
    
    # Fechando apenas a pasta
    excel.Workbooks.Close()
        
    workbook = load_workbook(filename=caminho_arquivo)
    
    os.system('cls')
    print(f'Calculando estatisticas...\n')
    
    # cria uma nova planilha com o nome "estatisticas"
    estatisticas_ws = workbook.create_sheet(title='estatisticas')
    
    media = None
    moda = None
    mediana = None
    maior_preco = None
    menor_preco = None
    
    col_d = list(dados_ws.iter_cols(min_row=2, min_col=4, max_col=4, max_row=dados_ws.max_row, values_only=True))[0]
    media = sum(col_d) / len(col_d)
    col_d_sorted = sorted(col_d)
    middle = len(col_d_sorted) // 2
    mediana = (col_d_sorted[middle] + col_d_sorted[-middle-1]) / 2
    maior_preco = max(col_d)
    menor_preco = min(col_d)
    
    # insere os valores na planilha "estatisticas"
    estatisticas_ws['A1'] = 'MÉDIA_PRECO'
    estatisticas_ws['A2'] = media
    estatisticas_ws['B1'] = 'MEDIANA_PRECO'
    estatisticas_ws['B2'] = mediana
    estatisticas_ws['C1'] = 'MENOR_PRECO'
    estatisticas_ws['C2'] = menor_preco
    estatisticas_ws['D1'] = 'MAIOR_PRECO'
    estatisticas_ws['D2'] = maior_preco

# Salvando o arquivo
workbook.save(f'{caminho_arquivo}')

# Abre o excel
excel = win32.gencache.EnsureDispatch('Excel.Application')
excel.Workbooks.Open(caminho_arquivo)
excel.Visible = True

os.system('cls')
print(f'Planilha reaberta!\nAba "estatisticas" criada!\n\nFIM!\n')